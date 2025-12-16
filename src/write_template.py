import pandas as pd
from openpyxl import load_workbook

def _header_map(ws):
    m = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            m[str(v).strip()] = c
    return m

def append_df_by_header(wb, sheet_name: str, df: pd.DataFrame):
    if df is None or df.empty:
        return
    ws = wb[sheet_name]
    hm = _header_map(ws)
    start_row = ws.max_row + 1
    for i, row in enumerate(df.to_dict(orient="records")):
        r = start_row + i
        for col_name, val in row.items():
            if col_name not in hm:
                continue
            ws.cell(row=r, column=hm[col_name], value=val)

def write_to_template(template_path: str, out_path: str, event_df: pd.DataFrame, evidence_df: pd.DataFrame, company_df: pd.DataFrame):
    wb = load_workbook(template_path)
    append_df_by_header(wb, "Company_Master", company_df)
    append_df_by_header(wb, "Event_Cards", event_df)
    append_df_by_header(wb, "Evidence_Log", evidence_df)
    wb.save(out_path)
